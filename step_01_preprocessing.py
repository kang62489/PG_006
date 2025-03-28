## Author: Kang
## Last Update: 2025-Mar-28
## Purpose: To analyze the specificity of the iAChSnFR and Integrate the old codes

# Modules
import os
import sys
import pandas as pd
import openpyxl
from pathlib import Path
from tabulate import tabulate
from rich import print
from classes.dialog_INQUIARY import INQUIRY
from PySide6.QtWidgets import QApplication

from functions.truncate_and_crop import truncate_and_crop

# Set event handler for GUIs (dialogs)
app = QApplication(sys.argv)

# File I/O
project_path = "E:/PRJ_iAChSnFR_Specificity/Data"
export_path = Path(__file__).parent / "Outputs"

# Dates of experiments
selected_folders = ["2023_10_05","2023_10_12","2024_03_07","2024_03_22"]

# Read summary csv files for getting all information of recording files of the selected folders
lst_loaded_summaries = []
for date in selected_folders:
    csv_filename = "{}_summary.csv".format(date)
    print(csv_filename)
    table_loaded = pd.read_csv(os.path.join(project_path, date, csv_filename))
    lst_loaded_summaries.append(table_loaded)
    print(tabulate(table_loaded, headers='keys', tablefmt='pretty'),"\n\n")
    
# Find the files fit the conditions in each date of folders
# Conditions:
# PUFF_PERIOD = 30 ms
# PUFF_COUNT = 1
# BATHED_IN = ACSF, NEO

df_all_selected_files_ACSF = pd.DataFrame()
df_all_selected_files_NEO = pd.DataFrame()
for csv_table in lst_loaded_summaries:
    filtered_table_ACSF = csv_table[(csv_table["PUFF_COUNT"]==1) & (csv_table["PUFF_PERIOD"]=="30ms") & (csv_table["BATHED_IN"] != 'NEO')]
    filtered_table_NEO = csv_table[(csv_table["PUFF_COUNT"]==1) & (csv_table["PUFF_PERIOD"]=="30ms") & (csv_table["BATHED_IN"] == 'NEO')]

    df_all_selected_files_ACSF = pd.concat([df_all_selected_files_ACSF, filtered_table_ACSF]).reset_index(drop=True)   
    df_all_selected_files_NEO = pd.concat([df_all_selected_files_NEO, filtered_table_NEO]).reset_index(drop=True)

print(tabulate(df_all_selected_files_ACSF, headers='keys', tablefmt='pretty'))
print(tabulate(df_all_selected_files_NEO, headers='keys', tablefmt='pretty'))

# Save the lists of selected files in xlsx format
export_xlsx_filename = export_path / "all_picked_files.xlsx"
if not os.path.exists(export_path):
    os.mkdir(export_path)
    
if not os.path.exists(export_path / export_xlsx_filename):    
    new_wb = openpyxl.Workbook()
    new_wb.save(export_xlsx_filename)
    
to_be_updated_sheets = ['Group_ACSF', 'Group_NEO']
to_be_updated_dataframes = {'Group_ACSF': df_all_selected_files_ACSF, 'Group_NEO': df_all_selected_files_NEO}

for sheet in to_be_updated_sheets:
    wb = openpyxl.load_workbook(export_xlsx_filename)
    if sheet in wb.sheetnames:
        del wb[sheet]
        wb.save(export_xlsx_filename)

    with pd.ExcelWriter(export_xlsx_filename,mode="a") as f:
        to_be_updated_dataframes[sheet].to_excel(f,sheet_name=sheet, index=False)

wb_final = openpyxl.load_workbook(export_xlsx_filename)
if "Sheet" in wb_final.sheetnames:
    del wb_final['Sheet']
    wb_final.save(export_xlsx_filename)    

# Truncate and crop the imaging files
# Load and trim the files
cut_start = 10
cut_end = 400
ROI_sizes = [512]
center_x = 1792
center_y = 1024

for r in ROI_sizes:
    output_foldername = ["Preprocessed_ACSF_"+str(r),"Preprocessed_NEO_"+str(r)]
    ROI_coords = [int(center_y-(r/2)), int(center_y+(r/2)), int(center_x-(r/2)), int(center_x+(r/2))]
    for outdir in output_foldername:
        # Check if preprocessing folder exists
        if os.path.exists(export_path / outdir):
            title = "Warning"
            message = f"Folder {outdir} already exists. Do you want to continue the process?"
            dlg_checkContinue = INQUIRY(title, message)
            if dlg_checkContinue.exec():
                truncate_and_crop(selected_folders, output_foldername, project_path, export_path, df_all_selected_files_ACSF, df_all_selected_files_NEO, ROI_coords, cut_start, cut_end, r)
            else:
                print(f"Preprocessing has been canceled or skipped.")
                break
        else:
            truncate_and_crop(selected_folders, output_foldername, project_path, export_path, df_all_selected_files_ACSF, df_all_selected_files_NEO, ROI_coords, cut_start, cut_end, r)
